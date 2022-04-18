# let's say we need to decrease the price by %10 -> =C2*0.9
# use the "as" keyword to give alias
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx")
# select a sheet. case-sensitive
sheet = wb["Sheet1"]
# give the coordinates of the cell u want to access. combination of column and row
# cell = sheet["a1"]
# OR
# cell = sheet.cell(1, 1)
# print(cell.value)

# we need to loop through over these rows and get the value of the third column which is the price and then multiply it by 0.9

# first we need to know how many rows are in the sheet
# print(sheet.max_row)

# we need create a for loop that will generate the numbers 1 to 4
# we don't want the first cell(price title), so we start from 2
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price
values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transaction2.xlsx')

